"""
Export all competitor data to Excel for Battlecard Data
Creates Competitor_Battlecard_Data.xlsx with all 115 fields for 82 competitors
"""

import os
import sys
from datetime import datetime

# Add backend to path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from database import SessionLocal, Competitor
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

def export_battlecard_data():
    """Export all competitor data to Excel with conditional formatting."""

    db = SessionLocal()

    try:
        # Get all active competitors
        competitors = db.query(Competitor).filter(
            Competitor.is_deleted == False
        ).order_by(Competitor.name).all()

        print(f"Found {len(competitors)} competitors")

        # Define all fields in order (115 fields)
        fields = [
            # Core Identity
            ('name', 'Company Name'),
            ('website', 'Website'),
            ('status', 'Status'),
            ('notes', 'Notes'),

            # Threat & Quality
            ('threat_level', 'Threat Level'),
            ('last_updated', 'Last Updated'),
            ('data_quality_score', 'Data Quality Score'),
            ('last_verified_at', 'Last Verified'),

            # Pricing
            ('pricing_model', 'Pricing Model'),
            ('base_price', 'Base Price'),
            ('price_unit', 'Price Unit'),

            # Product
            ('product_categories', 'Product Categories'),
            ('key_features', 'Key Features'),
            ('integration_partners', 'Integration Partners'),
            ('certifications', 'Certifications'),
            ('product_count', 'Product Count'),

            # Market & Customers
            ('target_segments', 'Target Segments'),
            ('customer_size_focus', 'Customer Size Focus'),
            ('geographic_focus', 'Geographic Focus'),
            ('customer_count', 'Customer Count'),
            ('customer_acquisition_rate', 'Customer Acquisition Rate'),
            ('key_customers', 'Key Customers'),
            ('g2_rating', 'G2 Rating'),

            # Company Info
            ('employee_count', 'Employee Count'),
            ('employee_growth_rate', 'Employee Growth Rate'),
            ('year_founded', 'Year Founded'),
            ('headquarters', 'Headquarters'),
            ('funding_total', 'Total Funding'),
            ('latest_round', 'Latest Round'),
            ('pe_vc_backers', 'PE/VC Backers'),

            # Digital Presence
            ('website_traffic', 'Website Traffic'),
            ('social_following', 'Social Following'),
            ('recent_launches', 'Recent Launches'),
            ('news_mentions', 'News Mentions'),

            # Public Company Data
            ('is_public', 'Is Public'),
            ('ticker_symbol', 'Ticker Symbol'),
            ('stock_exchange', 'Stock Exchange'),
            ('sec_cik', 'SEC CIK'),
            ('annual_revenue', 'Annual Revenue (SEC)'),
            ('net_income', 'Net Income (SEC)'),
            ('sec_employee_count', 'Employee Count (SEC)'),
            ('fiscal_year_end', 'Fiscal Year End'),
            ('recent_sec_filings', 'Recent SEC Filings'),
            ('sec_risk_factors', 'SEC Risk Factors'),

            # Market Vertical Tracking
            ('primary_market', 'Primary Market'),
            ('markets_served', 'Markets Served'),
            ('market_focus_score', 'Market Focus Score'),

            # Product Overlap
            ('has_pxp', 'Has PXP'),
            ('has_pms', 'Has PMS'),
            ('has_rcm', 'Has RCM'),
            ('has_patient_mgmt', 'Has Patient Mgmt'),
            ('has_payments', 'Has Payments'),
            ('has_biometric', 'Has Biometric'),
            ('has_interoperability', 'Has Interoperability'),
            ('product_overlap_score', 'Product Overlap Score'),

            # Enhanced Analytics
            ('telehealth_capabilities', 'Telehealth Capabilities'),
            ('ai_features', 'AI Features'),
            ('mobile_app_available', 'Mobile App Available'),
            ('hipaa_compliant', 'HIPAA Compliant'),
            ('ehr_integrations', 'EHR Integrations'),

            # Sales & Marketing Dimensions
            ('dim_product_packaging_score', 'Dim: Product Packaging Score'),
            ('dim_product_packaging_evidence', 'Dim: Product Packaging Evidence'),
            ('dim_integration_depth_score', 'Dim: Integration Depth Score'),
            ('dim_integration_depth_evidence', 'Dim: Integration Depth Evidence'),
            ('dim_support_service_score', 'Dim: Support Service Score'),
            ('dim_support_service_evidence', 'Dim: Support Service Evidence'),
            ('dim_retention_stickiness_score', 'Dim: Retention Stickiness Score'),
            ('dim_retention_stickiness_evidence', 'Dim: Retention Stickiness Evidence'),
            ('dim_user_adoption_score', 'Dim: User Adoption Score'),
            ('dim_user_adoption_evidence', 'Dim: User Adoption Evidence'),
            ('dim_implementation_ttv_score', 'Dim: Implementation TTV Score'),
            ('dim_implementation_ttv_evidence', 'Dim: Implementation TTV Evidence'),
            ('dim_reliability_enterprise_score', 'Dim: Reliability Enterprise Score'),
            ('dim_reliability_enterprise_evidence', 'Dim: Reliability Enterprise Evidence'),
            ('dim_pricing_flexibility_score', 'Dim: Pricing Flexibility Score'),
            ('dim_pricing_flexibility_evidence', 'Dim: Pricing Flexibility Evidence'),
            ('dim_reporting_analytics_score', 'Dim: Reporting Analytics Score'),
            ('dim_reporting_analytics_evidence', 'Dim: Reporting Analytics Evidence'),
            ('dim_overall_score', 'Dim: Overall Score'),
            ('dim_sales_priority', 'Dim: Sales Priority'),

            # Social Media Metrics
            ('linkedin_followers', 'LinkedIn Followers'),
            ('linkedin_employees', 'LinkedIn Employees'),
            ('linkedin_url', 'LinkedIn URL'),
            ('twitter_followers', 'Twitter Followers'),
            ('twitter_handle', 'Twitter Handle'),
            ('facebook_followers', 'Facebook Followers'),
            ('instagram_followers', 'Instagram Followers'),
            ('youtube_subscribers', 'YouTube Subscribers'),

            # Financial Metrics
            ('estimated_revenue', 'Estimated Revenue'),
            ('revenue_growth_rate', 'Revenue Growth Rate'),
            ('profit_margin', 'Profit Margin'),
            ('estimated_valuation', 'Estimated Valuation'),
            ('burn_rate', 'Burn Rate'),
            ('runway_months', 'Runway Months'),
            ('last_funding_date', 'Last Funding Date'),
            ('funding_stage', 'Funding Stage'),
            ('debt_financing', 'Debt Financing'),
            ('revenue_per_employee', 'Revenue Per Employee'),

            # Leadership & Team
            ('ceo_name', 'CEO Name'),
            ('ceo_linkedin', 'CEO LinkedIn'),
            ('cto_name', 'CTO Name'),
            ('cfo_name', 'CFO Name'),
            ('executive_changes', 'Executive Changes'),
            ('board_members', 'Board Members'),
            ('advisors', 'Advisors'),
            ('founder_background', 'Founder Background'),

            # Employee & Culture
            ('glassdoor_rating', 'Glassdoor Rating'),
            ('glassdoor_reviews_count', 'Glassdoor Reviews Count'),
            ('glassdoor_recommend_pct', 'Glassdoor Recommend %'),
            ('indeed_rating', 'Indeed Rating'),
            ('employee_turnover_rate', 'Employee Turnover Rate'),
            ('hiring_velocity', 'Hiring Velocity'),

            # Product & Technology
            ('latest_product_launch', 'Latest Product Launch'),
            ('tech_stack', 'Tech Stack'),
            ('cloud_provider', 'Cloud Provider'),
            ('api_available', 'API Available'),
            ('api_documentation_url', 'API Documentation URL'),
            ('open_source_contributions', 'Open Source Contributions'),
            ('rd_investment_pct', 'R&D Investment %'),

            # Market & Competitive
            ('estimated_market_share', 'Estimated Market Share'),
            ('nps_score', 'NPS Score'),
            ('customer_churn_rate', 'Customer Churn Rate'),
            ('average_contract_value', 'Average Contract Value'),
            ('sales_cycle_length', 'Sales Cycle Length'),
            ('competitive_win_rate', 'Competitive Win Rate'),

            # Regulatory & Compliance
            ('soc2_certified', 'SOC2 Certified'),
            ('hitrust_certified', 'HITRUST Certified'),
            ('iso27001_certified', 'ISO 27001 Certified'),
            ('legal_issues', 'Legal Issues'),

            # Patents & IP
            ('patent_count', 'Patent Count'),
            ('recent_patents', 'Recent Patents'),
            ('trademark_count', 'Trademark Count'),
            ('ip_litigation', 'IP Litigation'),

            # Partnerships & Ecosystem
            ('strategic_partners', 'Strategic Partners'),
            ('reseller_partners', 'Reseller Partners'),
            ('marketplace_presence', 'Marketplace Presence'),
            ('acquisition_history', 'Acquisition History'),

            # Customer Intelligence
            ('notable_customer_wins', 'Notable Customer Wins'),
            ('customer_case_studies', 'Customer Case Studies'),

            # External API Data
            ('logo_url', 'Logo URL'),
            ('email_pattern', 'Email Pattern'),
            ('key_contacts', 'Key Contacts'),
            ('hunter_email_count', 'Hunter Email Count'),

            # Metadata
            ('created_at', 'Created At'),
            ('extended_data_updated', 'Extended Data Updated'),
        ]

        print(f"Exporting {len(fields)} fields per competitor")

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Battlecard Data"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Row 1: Title
        ws.merge_cells('A1:E1')
        ws['A1'] = f"Competitor Battlecard Data Export - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws['A1'].font = Font(bold=True, size=14)

        # Row 2: Headers
        # Column A: Row number, Column B: Company Name, then all fields
        ws['A2'] = '#'
        ws['A2'].font = header_font
        ws['A2'].fill = header_fill
        ws['A2'].alignment = header_alignment
        ws['A2'].border = thin_border

        # Write field headers starting at column B
        for col_idx, (field_name, display_name) in enumerate(fields, start=2):
            cell = ws.cell(row=2, column=col_idx, value=display_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Write competitor data starting at row 3
        for row_idx, comp in enumerate(competitors, start=3):
            # Column A: Row number
            ws.cell(row=row_idx, column=1, value=row_idx - 2)

            # Write each field value
            for col_idx, (field_name, display_name) in enumerate(fields, start=2):
                value = getattr(comp, field_name, None)

                # Format the value
                if value is None or value == '' or value == []:
                    cell_value = 'N/A'
                elif isinstance(value, bool):
                    cell_value = 'Yes' if value else 'No'
                elif isinstance(value, datetime):
                    cell_value = value.strftime('%Y-%m-%d %H:%M')
                else:
                    cell_value = str(value)

                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        # Set column widths
        ws.column_dimensions['A'].width = 5  # Row number
        ws.column_dimensions['B'].width = 30  # Company name

        for col_idx in range(3, len(fields) + 2):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 20

        # Freeze panes (freeze row 2 and column B)
        ws.freeze_panes = 'C3'

        # Add conditional formatting for N/A cells (light red fill)
        light_red_fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")

        # Apply to all data cells (B3 to last column, last row)
        last_col_letter = get_column_letter(len(fields) + 1)
        last_row = len(competitors) + 2
        data_range = f"B3:{last_col_letter}{last_row}"

        ws.conditional_formatting.add(
            data_range,
            CellIsRule(
                operator='equal',
                formula=['"N/A"'],
                fill=light_red_fill
            )
        )

        # Also add conditional formatting for column A (though it won't have N/A)
        ws.conditional_formatting.add(
            f"A3:A{last_row}",
            CellIsRule(
                operator='equal',
                formula=['"N/A"'],
                fill=light_red_fill
            )
        )

        # Save the workbook
        output_path = os.path.join(os.path.dirname(__file__), '..', 'Competitor_Battlecard_Data.xlsx')
        wb.save(output_path)

        print(f"\nExcel file saved to: {output_path}")
        print(f"Total competitors: {len(competitors)}")
        print(f"Total fields: {len(fields)}")
        print(f"Total data cells: {len(competitors) * len(fields)}")

        # Count N/A values
        na_count = 0
        filled_count = 0
        for comp in competitors:
            for field_name, _ in fields:
                value = getattr(comp, field_name, None)
                if value is None or value == '' or value == []:
                    na_count += 1
                else:
                    filled_count += 1

        print(f"\nData coverage:")
        print(f"  Filled cells: {filled_count} ({filled_count / (len(competitors) * len(fields)) * 100:.1f}%)")
        print(f"  N/A cells: {na_count} ({na_count / (len(competitors) * len(fields)) * 100:.1f}%)")

        return output_path

    finally:
        db.close()


if __name__ == "__main__":
    export_battlecard_data()
